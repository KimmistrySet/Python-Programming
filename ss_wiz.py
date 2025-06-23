import openpyxl


def create_excel_file(filename, data=None):
    # Create a new Excel workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"

    # Adding headers to the first row
    headers = ["ID", "Name", "Age", "Email", "Phone"]
    worksheet.append(headers)
    # Set column widths for better readability
    column_widths = {
        "A": 10,  # ID
        "B": 20,  # Name
        "C": 10,  # Age
        "D": 30,  # Email
        "E": 15,  # Phone
    }
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    # Set the header row to bold
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    # Create a new sheet for data entry
    worksheet = workbook.create_sheet(title="Data Entry")
    # Set the headers for the data entry sheet
    data_entry_headers = ["Field1", "Field2"]
    worksheet.append(data_entry_headers)
    # Set column widths for the data entry sheet
    data_entry_column_widths = {"A": 20, "B": 20}  # Field1  # Field2
    for col, width in data_entry_column_widths.items():
        worksheet.column_dimensions[col].width = width
    # Set the header row to bold in the data entry sheet
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    # Create a new sheet for summary
    worksheet = workbook.create_sheet(title="Summary")
    # Set the headers for the summary sheet
    summary_headers = ["Summary Field1", "Summary Field2"]
    worksheet.append(summary_headers)
    # Set column widths for the summary sheet
    # Summary Field1  # Summary Field2
    summary_column_widths = {"A": 20, "B": 20}
    for col, width in summary_column_widths.items():
        worksheet.column_dimensions[col].width = width
    # Set the header row to bold in the summary sheet
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    # Create a new sheet for analysis
    worksheet = workbook.create_sheet(title="Analysis")
    # Set the headers for the analysis sheet
    analysis_headers = ["Analysis Field1", "Analysis Field2"]
    worksheet.append(analysis_headers)
    # Set column widths for the analysis sheet
    # Analysis Field1  # Analysis Field2
    analysis_column_widths = {"A": 20, "B": 20}
    for col, width in analysis_column_widths.items():
        worksheet.column_dimensions[col].width = width
    # Set the header row to bold in the analysis sheet
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    # Create a new sheet for results
    worksheet = workbook.create_sheet(title="Results")
    # Set the headers for the results sheet
    results_headers = ["Result Field1", "Result Field2"]
    worksheet.append(results_headers)
    # Set column widths for the results sheet
    results_column_widths = {
        "A": 20, "B": 20
    }  # Result Field1  # Result Field2
    for col, width in results_column_widths.items():
        worksheet.column_dimensions[col].width = width
    # Set the header row to bold in the results sheet
    for cell in worksheet[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    # Add some sample data
    worksheet["A1"] = "Hello"
    worksheet["B1"] = "World"
    worksheet["A2"] = 123
    worksheet["B2"] = 456
    # Add some sample data to the data entry sheet
    worksheet = workbook["Data Entry"]
    worksheet["A2"] = "Sample Data 1"
    worksheet["B2"] = "Sample Data 2"
    # Add some sample data to the summary sheet
    worksheet = workbook["Summary"]
    worksheet["A2"] = "Summary Data 1"
    worksheet["B2"] = "Summary Data 2"
    # Add some sample data to the analysis sheet
    worksheet = workbook["Analysis"]
    worksheet["A2"] = "Analysis Data 1"
    worksheet["B2"] = "Analysis Data 2"
    # Add some sample data to the results sheet
    worksheet = workbook["Results"]
    worksheet["A2"] = "Result Data 1"
    worksheet["B2"] = "Result Data 2"
    # Set the filename for the Excel file
    # filename = "sample_excel_file.xlsx"
    filename = filename if filename.endswith(".xlsx") else filename + ".xlsx"
    # Check if the filename is valid
    if not isinstance(filename, str) or not filename:
        raise ValueError(
            "Invalid file name provided. Please provide a valid file name."
        )
    # Check if the filename has a valid extension
    if not filename.endswith(".xlsx"):
        raise ValueError(
            (
                "Invalid file extension. Please provide a file name with "
                "'.xlsx' extension."
            )
        )
    # Check if the filename is too long
    if len(filename) > 255:
        raise ValueError(
            "File name is too long. Please provide a shorter file name."
        )
    # Check if the filename contains invalid characters
    invalid_chars = r'<>:"/\|?*'
    if any(char in filename for char in invalid_chars):
        raise ValueError(
            f"File name contains invalid characters: {invalid_chars}"
        )
    # Check if the filename is not empty
    if not filename.strip():
        raise ValueError(
            "File name cannot be empty. Please provide a valid file name."
        )
    # Check if the filename is not too short
    if len(filename) < 5:  # Assuming a minimum length of 5 characters
        raise ValueError(
            "File name is too short. Please provide a longer file name."
        )
    # Check if the filename does not contain only whitespace
    if filename.isspace():
        raise ValueError(
            "File name cannot contain only whitespace. "
            "Please provide a valid file name."
        )
    # Check if the filename does not contain only special characters
    if all(char in r'<>:"/\|?*' for char in filename):
        raise ValueError(
            "File name cannot contain only special characters. "
            "Please provide a valid file name."
        )
    # Check if the filename does not contain reserved words
    reserved_words = [
        "CON",
        "PRN",
        "AUX",
        "NUL",
        "COM1",
        "COM2",
        "COM3",
        "COM4",
        "COM5",
        "COM6",
        "COM7",
        "COM8",
        "COM9",
        "LPT1",
        "LPT2",
        "LPT3",
    ]
    if any(filename.upper().startswith(word) for word in reserved_words):
        raise ValueError(
            "File name cannot be a reserved word (e.g., CON, PRN, AUX, NUL, "
            "COM1, LPT1, etc.)."
        )
    for row in data:
        if not isinstance(row, list) or len(row) != len(headers):
            raise ValueError(
                "Each row of data must be a list with the same number of "
                "elements as the headers."
            )
        worksheet.append(row)
    # Save the workbook to the specified filename
    workbook.save(filename)
    print(f"Excel file '{filename}' successfully created with sample data.")


def update_excel_file(filename, data):
    # Load the existing workbook
    workbook = openpyxl.load_workbook(filename)
    # Select the active worksheet
    worksheet = workbook.active
    # Append new data to the worksheet
    for row in data:
        if not isinstance(row, list) or len(row) != worksheet.max_column:
            raise ValueError(
                "Each row of data must be a list with the same number of "
                "elements as the headers."
            )
        worksheet.append(row)
    # Save the updated workbook
    workbook.save(filename)
    print(f"Excel file '{filename}' successfully updated with new data.")


if __name__ == "__main__":
    # Example usage
    excel_filename = "sample_excel_file.xlsx"
    # Sample data to create the Excel file
    initial_data = [
        [1, "Alice", 30, "alice@example.com", "123-456-7890"],
        [2, "Bob", 25, "bob@example.com", "987-654-3210"],
        [3, "Charlie", 35, "charlie@example.com", "555-555-5555"],
    ]
    create_excel_file(excel_filename, data=initial_data)

    # Sample data to update the Excel file
    new_data = [
        [4, "David", 28, "david@example.com", "111-222-3333"],
        [5, "Eve", 22, "eve@example.com", "444-555-6666"],
    ]
    update_excel_file(excel_filename, new_data)
